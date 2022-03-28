"""Read all Excel files in current working directory and output them as CSV files.

Will make a CSV file for each sheet in each Excel workbook.

Functions:
    extensions_in_directory(path) -> iterable

TODO:
change it to operate outside of CWD
refactor to not have so many nested for loops and to have definitions/logic separated out more cleanly
use writerowS instead and just provide an iterable of rows (list of contents)
docstring for main function
"""

import csv
import logging
from pathlib import Path
from typing import Iterable

import openpyxl

logging.basicConfig(
    filename="excel-to-csv-converter_DEBUG.txt",
    level=logging.DEBUG,
    format=" %(asctime)s - %(levelname)s - %(message)s",
)
logging.disable(logging.CRITICAL)

def extensions_in_directory(directory: Path, extension: str) -> Iterable[Path]:
    """Return every file with the correct extension in directory"""
    yield from directory.glob(f"*{extension}")

def excel_to_csv():
    directory = Path.cwd()
    for excel_file in extensions_in_directory(directory, ".xlsx"):
        workbook = openpyxl.load_workbook(excel_file)
        for sheetname in workbook.sheetnames:
            csv_filename = f"{excel_file.stem}_{sheetname}.csv"
            sheet = workbook[sheetname]
            sheet_height = range(1, sheet.max_row + 1)
            sheet_width = range(1, sheet.max_column + 1)
            with open(csv_filename, "w") as csv_file:
                csv_writer = csv.writer(csv_file)
                for row_num in sheet_height:
                    csv_writer.writerow(
                        [
                            sheet.cell(row=row_num, column=col_num).value 
                            for col_num in sheet_width
                        ]
                    )

if __name__ == "__main__":
    excel_to_csv()
